# import fiona
import geopandas as gpd
import pandas as pd
import warnings
from scipy.stats import t  # Para la distribución T de Student
from openpyxl import load_workbook
from openpyxl.styles import Font  # Para estilos de fuente
import openpyxl  # Asegúrate de importar openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from functions.auxiliary import clean_sheet_name, delete_folder, compressed_files


def generate_result(file_path: str, boo_grouped: bool, list_ha: list, list_t_student: list, result_path: str, origin_type: str):
    """
    This function processes input data, performs various aggregations and calculations, and then saves the results to an Excel file. 
    It handles both spatial and non-spatial data, depending on the input file type. The results are saved for each 'BIOMA' or 'N_COBERT' group, 
    with additional statistical calculations and formatted output.

    Args:
        file_path (str): The path to the input file (either .gdb or .xlsx format).
        boo_grouped (bool): A boolean flag indicating whether the data should be grouped by 'BIOMA'.
        list_ha (list): A list of area values in hectares corresponding to each group.
        list_t_student (list): A list of t-student values for statistical calculations.
        result_path (str): The directory path where the result files will be saved.
        origin_type (str): The type of the input file ('gdb' for GeoDataFrame or 'xlsx' for Excel).

    Returns:
        Tuple[bool, str]: A tuple indicating whether the operation was successful (True/False) and a message. 
                          Returns a success message or error description.
    """
    
    delete_folder(result_path)

    # Ignorar todas las advertencias
    warnings.filterwarnings("ignore", category=UserWarning)

    DECIMALES = 6

    if origin_type == "gdb":
        # Inicializamos los DataFrames para cada capa
        df_punto_muestreo_flora = None
        df_muestreo_flora_fustal_tb = None

        df_punto_muestreo_flora = gpd.read_file(file_path, layer="PuntoMuestreoFlora")
        df_muestreo_flora_fustal_tb = gpd.read_file(file_path, layer="MuestreoFloraFustalTB")
    else:
        df_combined = pd.read_excel(file_path)

    if boo_grouped:
        if origin_type == "gdb":
            df_muestreo_ecosistema = gpd.read_file(file_path, layer="Ecosistema")
            # Realizar el merge asegurando conservar las coordenadas del GeoDataFrame
            df_combined = df_muestreo_flora_fustal_tb[['ID_MUEST', 'VOL_TOTAL']].merge(
                df_punto_muestreo_flora[['ID_MUEST', 'N_COBERT', 'AREA_UM_ha', 'geometry']],
                on='ID_MUEST',
                how='inner'
            )
        if len(df_combined)>0:
            if origin_type == "gdb":
                # Convertir el resultado nuevamente a GeoDataFrame
                df_combined = gpd.GeoDataFrame(df_combined, geometry='geometry')
                # Aseguramos que ambos DataFrames tienen el mismo sistema de coordenadas
                df_combined = df_combined.to_crs(df_muestreo_ecosistema.crs)
                # Realizamos un join espacial para buscar los shapes de df_combined que están dentro de los shapes de df_muestreo_ecosistema
                joined = gpd.sjoin(df_combined, df_muestreo_ecosistema, how="inner", predicate="within")
                if -1 in list_ha:
                    df_grouped = joined.groupby(['BIOMA', 'N_COBERT', 'ID_MUEST']).agg({
                        'VOL_TOTAL': 'sum',  # Ejemplo: suma de volúmenes
                        'AREA_UM_ha': 'mean'  # Ejemplo: promedio del área
                    }).reset_index()
                else:
                    df_grouped = joined.groupby(['BIOMA', 'N_COBERT', 'ID_MUEST']).agg({
                        'VOL_TOTAL': 'sum'  # Ejemplo: suma de volúmenes
                    }).reset_index()
            else:
                df_grouped = df_combined

            # print(f"list(df_grouped.columns): {list(df_grouped.columns)}")

            # Agrupar por 'BIOMA' en df_grouped
            grouped = df_grouped.groupby('BIOMA')
            # Crear una lista de DataFrames, uno para cada 'BIOMA'
            dataframes_por_bioma = [group.reset_index(drop=True) for _, group in grouped]
            list_dataframe = dataframes_por_bioma
        else:
            return False, "No hay registros en los cuales coincida ID_MUEST"
        
    else:
        if origin_type == "gdb":
            # Realizar el merge asegurando conservar las coordenadas del GeoDataFrame
            df_combined = df_muestreo_flora_fustal_tb[['ID_MUEST', 'VOL_TOTAL']].merge(
                df_punto_muestreo_flora[['ID_MUEST', 'N_COBERT', 'AREA_UM_ha']],
                on='ID_MUEST',
                how='inner'
            )
        if len(df_combined)>0:
            if -1 in list_ha:
                df_grouped = df_combined.groupby(['N_COBERT', 'ID_MUEST']).agg(
                VOL_TOTAL=('VOL_TOTAL', 'sum'),
                AREA_UM_ha=('AREA_UM_ha', 'first')
                ).reset_index()
            else:
                df_grouped = df_combined.groupby(['N_COBERT', 'ID_MUEST']).agg(
                VOL_TOTAL=('VOL_TOTAL', 'sum'),
                ).reset_index()
            list_dataframe = [df_grouped]
        else:
            return False, "No hay registros en los cuales coincida ID_MUEST"     
      
    if  len(df_grouped) > 0:
        for df_grouped in list_dataframe:
            # print(f"df_grouped.columns: {df_grouped.columns}")
            if 'BIOMA' in df_grouped.columns:
                result_path_final = f"{result_path}result_BIOMA_{df_grouped['BIOMA'][0]}.xlsx"
                df_grouped = df_grouped.drop('BIOMA', axis=1)
            else:
                result_path_final = f"{result_path}result.xlsx"
            # print(f"result_path_final: {result_path_final}")
            
            with pd.ExcelWriter(result_path_final, engine='openpyxl') as writer:
                # Iterar sobre cada grupo de 'N_COBERT'
                for idx, (n_cobert, group_df) in enumerate(df_grouped.groupby('N_COBERT')):
                    group_df = group_df.drop(columns='N_COBERT')
                    # group_df['tmp'] = 1
                    # group_df['VOL_TOTAL'] = round(group_df['VOL_TOTAL'] / group_df['tmp'], DECIMALES)
                    # group_df = group_df.drop(columns='tmp')
                    if -1 in list_ha:
                        group_df['ha'] = group_df['AREA_UM_ha']
                        group_df = group_df.drop(columns=['AREA_UM_ha'])
                    else:
                        group_df['ha'] = list_ha[idx]

                    group_df['TOTAL'] = round(group_df['VOL_TOTAL'] / group_df['ha'], DECIMALES)

                    # print(f"group_df:\n {group_df}")
                    
                    total_row = pd.DataFrame({'ID_MUEST': ['Total'], 'VOL_TOTAL': [''], 'ha': [''], 
                                            'TOTAL': [group_df['TOTAL'].sum()]})
                    
                    
                    # Crear una fila en blanco
                    empty_row = pd.DataFrame({'ID_MUEST': [''], 'VOL_TOTAL': [''], 
                                            'ha': [''], 'TOTAL': ['']})

                    estadigrafo_row = pd.DataFrame({'ID_MUEST': ['ESTADIGRAFO'], 'VOL_TOTAL': [''], 
                                                    'ha': [''], 'TOTAL': ['VALOR']})

                    std_dev = round(group_df['TOTAL'].std(),DECIMALES)
                    mean_total = round(group_df['TOTAL'].mean(),DECIMALES)
                    num_parcelas = round(group_df['ID_MUEST'].nunique(),DECIMALES)
                    cv = round((std_dev / mean_total) * 100,DECIMALES)
                    error_std = round(std_dev / (num_parcelas ** 0.5),DECIMALES)
                    if -1 in list_t_student:
                        t_student = round(t.ppf(1 - 0.05, df=num_parcelas-1),DECIMALES)
                    else:
                        t_student = list_t_student[idx]
                    sx_t = round(error_std * t_student,DECIMALES)
                    percent_error = round((sx_t / mean_total) * 100,DECIMALES)

                    std_row = pd.DataFrame({'ID_MUEST': ['Desviación estándar (Ds)'], 'VOL_TOTAL': [''], 
                                            'ha': [''], 'TOTAL': [std_dev]})

                    media_row = pd.DataFrame({'ID_MUEST': ['Media (X)'], 'VOL_TOTAL': [''], 
                                            'ha': [''], 'TOTAL': [mean_total]})

                    num_parcelas_row = pd.DataFrame({'ID_MUEST': ['Número de parcelas'], 'VOL_TOTAL': [''], 
                                                    'ha': [''], 'TOTAL': [num_parcelas]})

                    cv_row = pd.DataFrame({'ID_MUEST': ['Coeficiente de variación (Cv)'], 'VOL_TOTAL': [''], 
                                        'ha': [''], 'TOTAL': [cv]})

                    sx_row = pd.DataFrame({'ID_MUEST': ['Error estándar (Sx)'], 'VOL_TOTAL': [''], 
                                        'ha': [''], 'TOTAL': [error_std]})

                    t_student_row = pd.DataFrame({'ID_MUEST': ['T (Student)'], 'VOL_TOTAL': [''], 
                                                'ha': [''], 'TOTAL': [t_student]})

                    sx_t_row = pd.DataFrame({'ID_MUEST': ['Sx * T'], 'VOL_TOTAL': [''], 
                                            'ha': [''], 'TOTAL': [sx_t]})

                    percent_error_row = pd.DataFrame({'ID_MUEST': ['% Error (Sx/t)/X'], 'VOL_TOTAL': [''], 
                                                    'ha': [''], 'TOTAL': [percent_error]})

                    cumple_row = pd.DataFrame({'ID_MUEST': ['Verificación Cumplimiento'], 'VOL_TOTAL': [''], 
                                            'ha': [''], 'TOTAL': ['SI CUMPLE' if percent_error < 15 else 'NO CUMPLE']})
                    
                    group_df_final = pd.concat([group_df, total_row, empty_row, estadigrafo_row, std_row, media_row, num_parcelas_row, 
                                        cv_row, sx_row, t_student_row, sx_t_row, percent_error_row, cumple_row], ignore_index=True)
                    
                    # Renombrar la primera columna a "PARCELA"
                    group_df_final.rename(columns={'ID_MUEST': 'PARCELA'}, inplace=True)

                    n_covert_clean = clean_sheet_name(n_cobert)

                    # Guardar cada DataFrame en una pestaña de Excel
                    group_df_final.to_excel(writer, sheet_name=str(n_covert_clean), index=False)

                    # Ajustar automáticamente el ancho de las columnas
                    ws = writer.sheets[str(n_covert_clean)]  # Obtener la hoja correspondiente
                    for column in group_df.columns:
                        max_length = max(group_df[column].astype(str).map(len).max(), len(column)) + 2  # +2 para un pequeño margen
                        col_letter = openpyxl.utils.get_column_letter(group_df.columns.get_loc(column) + 1)  # Obtener la letra de la columna
                        ws.column_dimensions[col_letter].width = max_length

            # Aplicar formato de negritas a la fila "Total" para cada hoja
            wb = load_workbook(result_path_final)

            # Definir los rellenos (fills) de colores
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # Rojo
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # Verde

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    if row[0].value == "Total":
                        for cell in row:
                            cell.font = Font(bold=True)
                    if row[0].value == "ESTADIGRAFO":
                        for cell in row:
                            cell.font = Font(bold=True)
                    if row[0].value == "% Error (Sx/t)/X":
                        for cell in row:
                            cell.font = Font(bold=True)
                    
                    # Aplicar color a la columna 'TOTAL'
                    total_column_index = ws.max_column
                    if row[total_column_index - 1].value == 'SI CUMPLE':
                        row[total_column_index - 1].fill = green_fill
                        row[total_column_index - 1].alignment = Alignment(horizontal='center')
                    elif row[total_column_index - 1].value == 'NO CUMPLE':
                        row[total_column_index - 1].fill = red_fill
                        row[total_column_index - 1].alignment = Alignment(horizontal='center')

            # Guardar los cambios en el archivo
            wb.save(result_path_final)

        compressed_files(result_path,"result.zip")
        return True, "ok"
    else:
        return False, "Valores nulos en N_COBERT o ID_MUEST"


